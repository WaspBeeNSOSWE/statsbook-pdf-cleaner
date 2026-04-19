#!/usr/bin/env python3
"""Statsbook PDF Cleaner.

Extracts only the sheets you actually need to print from a WFTDA StatsBook
exported by CRG scoreboard, strips all headers and footers, and saves a
clean PDF.

Usage:
    Place this script next to your CRG scoreboard folder (e.g. crg-scoreboard_v2025.8)
    and run it. It auto-detects the newest CRG version available.

    python3 statsbook-print.py
    python3 statsbook-print.py --crg /path/to/crg-scoreboard_v2025.8

Environment variables:
    CRG_PATH    Override path to the CRG scoreboard folder.

Requirements:
    pip install openpyxl pypdf
    LibreOffice installed (https://www.libreoffice.org/)
"""
import sys
import os
import shutil
import subprocess
import tempfile
import argparse
from pathlib import Path

# Sheets that should be kept in the printable PDF. Everything else is removed.
KEEP_SHEETS = {
    "IGRF",
    "Score",
    "Penalties",
    "Penalties-Lineups",
    "Expulsion-Suspension Form",
    "Official Reviews",
    "Penalty Box",
    "Rosters",
}


def find_soffice() -> str | None:
    """Find the LibreOffice executable across OSes."""
    candidates = [
        "soffice",
        "/Applications/LibreOffice.app/Contents/MacOS/soffice",
        r"C:\Program Files\LibreOffice\program\soffice.exe",
        r"C:\Program Files (x86)\LibreOffice\program\soffice.exe",
        "/usr/bin/soffice",
        "/usr/local/bin/soffice",
    ]
    for candidate in candidates:
        try:
            subprocess.run([candidate, "--version"], capture_output=True, timeout=10)
            return candidate
        except (FileNotFoundError, subprocess.TimeoutExpired):
            continue
    return None


def find_crg_folder(script_dir: Path, override: str | None = None) -> Path | None:
    """Find the CRG scoreboard xlsx export folder.

    Precedence:
    1. --crg command-line argument (override)
    2. CRG_PATH environment variable
    3. Auto-detect newest crg-scoreboard_v*/html/game-data/xlsx next to the script
    """
    if override:
        p = Path(override).expanduser().resolve()
        candidates = [p, p / "html" / "game-data" / "xlsx"]
        for c in candidates:
            if c.is_dir():
                return c
        return None

    env_path = os.environ.get("CRG_PATH")
    if env_path:
        p = Path(env_path).expanduser().resolve()
        candidates = [p, p / "html" / "game-data" / "xlsx"]
        for c in candidates:
            if c.is_dir():
                return c

    # Auto-detect: look for crg-scoreboard_v*/html/game-data/xlsx
    matches = []
    for crg_root in script_dir.glob("crg-scoreboard_v*"):
        xlsx_dir = crg_root / "html" / "game-data" / "xlsx"
        if xlsx_dir.is_dir():
            matches.append((crg_root, xlsx_dir))

    if not matches:
        return None

    # Use the newest (by modification time of the root)
    matches.sort(key=lambda x: x[0].stat().st_mtime, reverse=True)
    print(f"Using CRG scoreboard: {matches[0][0].name}")
    return matches[0][1]


def clean_and_convert(xlsx_path: Path, output_folder: Path, soffice: str) -> bool:
    """Clean a single statsbook file and save it as PDF. Returns True if written."""
    output_pdf = output_folder / f"{xlsx_path.stem}_PRINT.pdf"
    if output_pdf.exists() and output_pdf.stat().st_mtime >= xlsx_path.stat().st_mtime:
        return False

    print(f"Processing: {xlsx_path.name}")

    import openpyxl
    from pypdf import PdfReader

    with tempfile.TemporaryDirectory() as tmpdir:
        tmp_xlsx = Path(tmpdir) / xlsx_path.name
        shutil.copy2(xlsx_path, tmp_xlsx)

        wb = openpyxl.load_workbook(tmp_xlsx)

        # Remove sheets that shouldn't be printed
        for name in list(wb.sheetnames):
            if name not in KEEP_SHEETS:
                del wb[name]

        # Clear headers/footers — keep only the sheet name at the top (&A)
        for ws in wb.worksheets:
            ws.oddHeader.left.text = "&A"
            ws.oddHeader.center.text = None
            ws.oddHeader.right.text = None
            ws.oddFooter.left.text = None
            ws.oddFooter.center.text = None
            ws.oddFooter.right.text = None
            ws.evenHeader.left.text = None
            ws.evenHeader.center.text = None
            ws.evenHeader.right.text = None
            ws.evenFooter.left.text = None
            ws.evenFooter.center.text = None
            ws.evenFooter.right.text = None

        wb.save(tmp_xlsx)
        print(f"  Sheets kept: {', '.join(wb.sheetnames)}")

        # Convert to PDF via LibreOffice headless
        result = subprocess.run(
            [soffice, "--headless", "--convert-to", "pdf", "--outdir", tmpdir, str(tmp_xlsx)],
            capture_output=True,
            text=True,
            timeout=120,
        )

        tmp_pdf = Path(tmpdir) / f"{xlsx_path.stem}.pdf"
        if not tmp_pdf.exists():
            print("  ERROR: PDF conversion failed")
            if result.stderr:
                print(f"  {result.stderr.strip()}")
            return False

        shutil.move(tmp_pdf, output_pdf)
        pages = len(PdfReader(output_pdf).pages)
        print(f"  → {output_pdf.name} ({pages} pages)")
        return True


def main() -> int:
    parser = argparse.ArgumentParser(
        description="Clean and convert WFTDA StatsBook files to printable PDFs."
    )
    parser.add_argument(
        "--crg",
        help="Path to CRG scoreboard folder or its xlsx subfolder. "
        "Overrides auto-detection and CRG_PATH.",
    )
    parser.add_argument(
        "--output",
        default=None,
        help="Output folder for PDFs. Default: Statsbooks/ next to this script.",
    )
    args = parser.parse_args()

    script_dir = Path(__file__).resolve().parent
    output_folder = Path(args.output).expanduser().resolve() if args.output else script_dir / "Statsbooks"
    output_folder.mkdir(exist_ok=True)

    try:
        import openpyxl  # noqa: F401
        from pypdf import PdfReader  # noqa: F401
    except ImportError as e:
        print(f"ERROR: Missing Python package: {e}")
        print("Run: pip install openpyxl pypdf")
        return 1

    soffice = find_soffice()
    if not soffice:
        print("ERROR: LibreOffice (soffice) not found.")
        print("Install LibreOffice from https://www.libreoffice.org/")
        return 1

    crg_folder = find_crg_folder(script_dir, override=args.crg)
    if not crg_folder:
        print("ERROR: Could not locate a CRG scoreboard folder.")
        print("Expected: a crg-scoreboard_v*/ folder next to this script,")
        print("or pass --crg /path/to/crg-scoreboard_vX.Y")
        print("or set the CRG_PATH environment variable.")
        return 1

    xlsx_files = sorted(crg_folder.glob("STATS-*.xlsx"), key=os.path.getmtime, reverse=True)
    if not xlsx_files:
        print(f"No statsbooks found in {crg_folder}")
        print("Click 'Update' in CRG to export the current game, then try again.")
        return 1

    processed = 0
    for xlsx_path in xlsx_files:
        if clean_and_convert(xlsx_path, output_folder, soffice):
            processed += 1

    if processed == 0:
        print("\nAll games already processed.")
    else:
        print(f"\nDone! {processed} game(s) processed.")
    return 0


if __name__ == "__main__":
    sys.exit(main())
